import openpyxl

# load the Excel
excelfile = openpyxl.load_workbook("inventory.xlsx")

# go to specific tab or sheet
pdt_list = excelfile["Sheet1"]

# products per supplier
pdts_per_sup = {}

total_value_per_sup = {}

# number of rows
print(pdt_list.max_row)

# range() create a list 0 to 74
# for i in 75 is not correct
# excel contains row from 2 to 75
for pdt_row in range(2, pdt_list.max_row + 1):
    sup_name = pdt_list.cell(pdt_row, 4).value
    inventory = pdt_list.cell(pdt_row, 2).value
    price = pdt_list.cell(pdt_row, 3).value
    # adding a new column in excel
    inventory_price = pdt_list.cell(pdt_row, 5)

    if sup_name in pdts_per_sup:
        pdts_per_sup[sup_name] = pdts_per_sup[sup_name] + 1
        # pdts_per_sup[sup_name] = pdts_per_sup.get(sup_name) + 1
    else:
        print("adding new supplier")
        pdts_per_sup[sup_name] = 1

    if sup_name in total_value_per_sup:
        total_value_per_sup[sup_name] = total_value_per_sup[sup_name] + (inventory * price)
    else:
        print("adding new supplier for inventory")
        total_value_per_sup[sup_name] = inventory * price

    inventory_price.value = inventory * price

excelfile.save("newexcelfile.xlsx")

print(f"pdts per supplier{pdts_per_sup}")
print(f"total value per supplier {total_value_per_sup}")
